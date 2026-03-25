"""Reconciliation logic — ported from Reconciliation_sources.

Uses openpyxl (already installed). All loaders accept bytes from HTTP uploads.
Supports Product (range) reconciliation and Vendor/Customer Invoice + OS reconciliation.
"""

from __future__ import annotations

import io
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _wb(file_bytes: bytes):
    return load_workbook(io.BytesIO(file_bytes), data_only=True)


def _clean_code(value) -> str | None:
    """Normalize a product code: strip whitespace and remove trailing .0 from floats."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
        return s
    except (ValueError, TypeError):
        return s


def _os_code(value) -> str | None:
    """Normalize an OS code: preserve/add leading zeros for 4-digit codes (e.g. 5 → '0005')."""
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        n = int(float(value))
        return f"{n:04d}" if 0 <= n < 10_000 else str(n)
    s = str(value).strip()
    if not s:
        return None
    if s.isdigit():
        n = int(s)
        if 0 <= n < 10_000:
            return f"{n:04d}"
    return s


def _stibo_col_idx(ws, aliases: tuple[str, ...]) -> int:
    """Return 1-based column index matching one of the normalized aliases, else 1."""
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        n = str(cell.value).strip().lower().replace(" ", "")
        if n in aliases:
            return idx
    return 1


# ---------------------------------------------------------------------------
# Product loaders
# ---------------------------------------------------------------------------


def load_jeeves_product(file_bytes: bytes) -> list[str]:
    """JEEVES: sheet '2-EXCELMASTER', column A from row 3."""
    wb = _wb(file_bytes)
    ws = wb["2-EXCELMASTER"]
    codes = []
    for (val,) in ws.iter_rows(min_row=3, min_col=1, max_col=1, values_only=True):
        c = _clean_code(val)
        if c:
            codes.append(c)
    return codes


def load_prophet_product(file_bytes: bytes) -> list[str]:
    """Prophet ERP: row 2 headers, 'FD Product Code' column, data from row 3."""
    wb = _wb(file_bytes)
    ws = wb.worksheets[0]
    headers = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1)}
    col_idx = headers.get("FD Product Code")
    if col_idx is None:
        available = [h for h in headers if h]
        raise ValueError(
            f"Column 'FD Product Code' not found in Prophet file. Available: {available}"
        )
    codes = []
    for row in range(3, ws.max_row + 1):
        c = _clean_code(ws.cell(row, col_idx).value)
        if c:
            codes.append(c)
    return codes


def load_ct_product(file_bytes: bytes, filename: str = "") -> list[str]:
    """CT Ekofisk: headers row 6, data from B7. Supports .xlsb and .xlsx."""
    if filename.lower().endswith(".xlsb"):
        return _load_ct_product_xlsb(file_bytes)
    wb = _wb(file_bytes)
    ws = next(
        (wb[n] for n in wb.sheetnames if "product" in n.lower()),
        wb.active,
    )
    codes = []
    for row in ws.iter_rows(min_row=7, min_col=2, max_col=2, values_only=True):
        c = _clean_code(row[0])
        if c:
            codes.append(c)
    return codes


def _load_ct_product_xlsb(file_bytes: bytes) -> list[str]:
    try:
        from pyxlsb import open_workbook
    except ModuleNotFoundError:
        raise ModuleNotFoundError(
            "pyxlsb is required for .xlsb files. Install with: pip install pyxlsb"
        )
    with open_workbook(io.BytesIO(file_bytes)) as wb:
        sheet_name = next((n for n in wb.sheets if n.lower() == "item"), wb.sheets[0])
        with wb.get_sheet(sheet_name) as sheet:
            rows = list(sheet.rows())
    codes = []
    for row_idx in range(6, len(rows)):
        row = rows[row_idx]
        if len(row) > 1:
            c = _clean_code(row[1].v)
            if c:
                codes.append(c)
    return codes


def load_stibo_product(file_bytes: bytes) -> list[str]:
    """STIBO: headers row 1, SUPC column from row 2."""
    wb = _wb(file_bytes)
    ws = wb.active
    supc_col = next(
        (
            idx
            for idx, cell in enumerate(ws[1], start=1)
            if cell.value and str(cell.value).strip().upper() == "SUPC"
        ),
        3,
    )
    codes = []
    for (val,) in ws.iter_rows(
        min_row=2, min_col=supc_col, max_col=supc_col, values_only=True
    ):
        c = _clean_code(val)
        if c:
            codes.append(c)
    return codes


def load_erp_product(file_bytes: bytes, erp_type: str, filename: str = "") -> list[str]:
    """Dispatch to the correct ERP product loader."""
    name = erp_type.strip().lower()
    if name == "jeeves":
        return load_jeeves_product(file_bytes)
    if name == "prophet":
        return load_prophet_product(file_bytes)
    raise ValueError(f"Unknown ERP type '{erp_type}'. Supported: Jeeves, Prophet.")


# ---------------------------------------------------------------------------
# Product reconciliation
# ---------------------------------------------------------------------------


def reconcile_products(
    ct_codes: list[str],
    erp_codes: list[str],
    stibo_codes: list[str],
    erp_name: str = "ERP",
) -> dict:
    """Build product range reconciliation table with presence marks and metrics."""
    ct_set = set(ct_codes)
    erp_set = set(erp_codes)
    stibo_set = set(stibo_codes)
    all_codes = sorted(ct_set | erp_set | stibo_set)

    rows = []
    for code in all_codes:
        in_ct = code in ct_set
        in_erp = code in erp_set
        in_stibo = code in stibo_set
        absent = [
            lbl
            for lbl, p in [("CT", in_ct), (erp_name, in_erp), ("STIBO", in_stibo)]
            if not p
        ]
        rows.append(
            {
                "ProductCode": code,
                "CT": "X" if in_ct else "",
                erp_name: "X" if in_erp else "",
                "STIBO": "X" if in_stibo else "",
                "Absent_from": ", ".join(absent) if absent else "-",
            }
        )

    metrics = {
        "total": len(rows),
        "in_all_3": sum(1 for r in rows if r["Absent_from"] == "-"),
        "ct_count": len(ct_set),
        "erp_count": len(erp_set),
        "stibo_count": len(stibo_set),
        "missing_ct": sum(1 for r in rows if not r["CT"]),
        "missing_erp": sum(1 for r in rows if not r[erp_name]),
        "missing_stibo": sum(1 for r in rows if not r["STIBO"]),
    }
    return {"rows": rows, "metrics": metrics, "erp_name": erp_name}


# ---------------------------------------------------------------------------
# Invoice / OS loaders — CT
# ---------------------------------------------------------------------------

_CT_FIRST_ROW = 8


def _load_ct_column(file_bytes: bytes, sheet_name: str, col: int) -> list[str]:
    wb = _wb(file_bytes)
    if sheet_name not in wb.sheetnames:
        available = wb.sheetnames
        raise ValueError(
            f"Sheet '{sheet_name}' not found in CT file. Available: {available}"
        )
    ws = wb[sheet_name]
    values = []
    for row in range(_CT_FIRST_ROW, ws.max_row + 1):
        v = ws.cell(row=row, column=col).value
        if v is None or (isinstance(v, str) and not v.strip()):
            break
        values.append(str(v).strip())
    return values


def load_ct_vendor_invoice(file_bytes: bytes) -> list[str]:
    return _load_ct_column(file_bytes, "Invoice", col=3)


def load_ct_vendor_os(file_bytes: bytes) -> list[str]:
    return _load_ct_column(file_bytes, "OrderingShipping", col=4)


def load_ct_customer_invoice(file_bytes: bytes) -> list[str]:
    return _load_ct_column(file_bytes, "Invoice", col=3)


def load_ct_customer_os(file_bytes: bytes) -> list[str]:
    return _load_ct_column(file_bytes, "OrderingShipping", col=4)


# ---------------------------------------------------------------------------
# Invoice / OS loaders — STIBO
# ---------------------------------------------------------------------------


def load_stibo_vendor_invoice(file_bytes: bytes) -> list[str]:
    """STIBO vendor invoice — auto-detects column by header or sheet name."""
    wb = _wb(file_bytes)
    # Extract format: sheet named "Invoice"
    if "Invoice" in wb.sheetnames:
        ws = wb["Invoice"]
        return [
            str(ws.cell(r, 1).value).strip()
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, 1).value is not None and str(ws.cell(r, 1).value).strip()
        ]
    ws = wb.active
    col = _stibo_col_idx(ws, ("suvcinvoice", "suvc-invoice"))
    return [
        str(ws.cell(r, col).value).strip()
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, col).value is not None and str(ws.cell(r, col).value).strip()
    ]


def load_stibo_vendor_os(file_bytes: bytes) -> list[str]:
    """STIBO vendor OS — auto-detects column by header or sheet name."""
    wb = _wb(file_bytes)
    if "Ordering-Shipping" in wb.sheetnames:
        ws = wb["Ordering-Shipping"]
        return [
            str(ws.cell(r, 1).value).strip()
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, 1).value is not None and str(ws.cell(r, 1).value).strip()
        ]
    ws = wb.active
    col = _stibo_col_idx(ws, ("suvcordering/shipping",))
    return [
        str(ws.cell(r, col).value).strip()
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, col).value is not None and str(ws.cell(r, col).value).strip()
    ]


def load_stibo_customer_invoice(file_bytes: bytes) -> list[str]:
    """STIBO customer invoice — detects 'Invoice Customer Code' column or sheet."""
    wb = _wb(file_bytes)
    if "Invoice" in wb.sheetnames:
        ws = wb["Invoice"]
        return [
            str(ws.cell(r, 1).value).strip()
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, 1).value is not None and str(ws.cell(r, 1).value).strip()
        ]
    ws = wb.active
    col = _stibo_col_idx(ws, ("invoicecustomercode",))
    return [
        str(ws.cell(r, col).value).strip()
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, col).value is not None and str(ws.cell(r, col).value).strip()
    ]


def load_stibo_customer_os(file_bytes: bytes) -> list[str]:
    """STIBO customer OS — detects OS column or sheet."""
    wb = _wb(file_bytes)
    if "Ordering-Shipping" in wb.sheetnames:
        ws = wb["Ordering-Shipping"]
        return [
            str(ws.cell(r, 1).value).strip()
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, 1).value is not None and str(ws.cell(r, 1).value).strip()
        ]
    ws = wb.active
    col = _stibo_col_idx(ws, ("suvcordering/shipping", "customerordershipping"))
    return [
        str(ws.cell(r, col).value).strip()
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, col).value is not None and str(ws.cell(r, col).value).strip()
    ]


# ---------------------------------------------------------------------------
# Invoice / OS loaders — Jeeves ERP
# ---------------------------------------------------------------------------

_JEEVES_VENDOR_OS_SHEETS = (
    "ORDERSHIPPING",
    "ODERSHIPPING",
    "OrderingShipping",
    "ORDERINGSHIPPING",
)
_JEEVES_CUSTOMER_INVOICE_SHEET = "INVOICECUSTOMER"
_JEEVES_CUSTOMER_OS_SHEETS = ("ORDERSHIPPING", "OrderShipping", "ORDERINGSHIPPING")


def load_jeeves_vendor_invoice(file_bytes: bytes) -> list[str]:
    """JEEVES Vendor: headers row 1, 'SUVC -Invoice' column, data from row 2."""
    wb = _wb(file_bytes)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_idx = next(
        (
            i + 1
            for i, h in enumerate(headers)
            if h and str(h).strip().lower().replace(" ", "") == "suvc-invoice"
        ),
        None,
    )
    if col_idx is None:
        raise ValueError(
            f"Column 'SUVC -Invoice' not found in Jeeves Vendor file. Headers: {headers}"
        )
    return [
        v
        for r in range(2, ws.max_row + 1)
        if (v := ws.cell(r, col_idx).value) is not None and str(v).strip()
    ]


def load_jeeves_vendor_os(file_bytes: bytes) -> list[str]:
    """JEEVES Vendor OS: sheet 'ORDERSHIPPING', column A from row 2."""
    wb = _wb(file_bytes)
    ws = next((wb[n] for n in _JEEVES_VENDOR_OS_SHEETS if n in wb.sheetnames), None)
    if ws is None:
        raise ValueError(
            f"Vendor OS sheet not found. Tried: {_JEEVES_VENDOR_OS_SHEETS}. "
            f"Available: {wb.sheetnames}"
        )
    return [
        str(ws.cell(r, 1).value).strip()
        for r in range(2, ws.max_row + 1)
        if ws.cell(r, 1).value is not None and str(ws.cell(r, 1).value).strip()
    ]


def load_jeeves_customer_invoice(file_bytes: bytes) -> list[str]:
    """JEEVES Customer Invoice: sheet 'INVOICECUSTOMER', column A from row 3."""
    wb = _wb(file_bytes)
    if _JEEVES_CUSTOMER_INVOICE_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{_JEEVES_CUSTOMER_INVOICE_SHEET}' not found. "
            f"Available: {wb.sheetnames}"
        )
    ws = wb[_JEEVES_CUSTOMER_INVOICE_SHEET]
    return [
        str(ws.cell(r, 1).value).strip()
        for r in range(3, ws.max_row + 1)
        if ws.cell(r, 1).value is not None and str(ws.cell(r, 1).value).strip()
    ]


def load_jeeves_customer_os(file_bytes: bytes) -> list[str]:
    """JEEVES Customer OS: column A from row 3. Preserves leading zeros (e.g. 5 → '0005')."""
    wb = _wb(file_bytes)
    ws = next((wb[n] for n in _JEEVES_CUSTOMER_OS_SHEETS if n in wb.sheetnames), None)
    if ws is None:
        raise ValueError(
            f"Customer OS sheet not found. Tried: {_JEEVES_CUSTOMER_OS_SHEETS}. "
            f"Available: {wb.sheetnames}"
        )
    codes = []
    for r in range(3, ws.max_row + 1):
        c = _os_code(ws.cell(r, 1).value)
        if c:
            codes.append(c)
    return codes


# ---------------------------------------------------------------------------
# ERP dispatch for Invoice / OS
# ---------------------------------------------------------------------------


def _load_erp_vendor_invoice(file_bytes: bytes, erp_name: str) -> list[str]:
    if erp_name.lower() == "jeeves":
        return load_jeeves_vendor_invoice(file_bytes)
    raise NotImplementedError(
        f"No Vendor Invoice loader implemented for ERP '{erp_name}'."
    )


def _load_erp_vendor_os(file_bytes: bytes, erp_name: str) -> list[str]:
    if erp_name.lower() == "jeeves":
        return load_jeeves_vendor_os(file_bytes)
    raise NotImplementedError(f"No Vendor OS loader implemented for ERP '{erp_name}'.")


def _load_erp_customer_invoice(file_bytes: bytes, erp_name: str) -> list[str]:
    if erp_name.lower() == "jeeves":
        return load_jeeves_customer_invoice(file_bytes)
    raise NotImplementedError(
        f"No Customer Invoice loader implemented for ERP '{erp_name}'."
    )


def _load_erp_customer_os(file_bytes: bytes, erp_name: str) -> list[str]:
    if erp_name.lower() == "jeeves":
        return load_jeeves_customer_os(file_bytes)
    raise NotImplementedError(
        f"No Customer OS loader implemented for ERP '{erp_name}'."
    )


# ---------------------------------------------------------------------------
# Invoice / OS reconciliation
# ---------------------------------------------------------------------------


def _normalize_os(codes: list) -> list[str]:
    result = []
    for v in codes:
        c = _os_code(v)
        if c:
            result.append(c)
    return result


def _build_entity_table(
    stibo: list[str],
    ct: list[str],
    erp: list[str],
    erp_name: str,
) -> dict:
    """Build a single-entity (vendor or customer) reconciliation view."""
    s_set, c_set, e_set = set(stibo), set(ct), set(erp)
    all_codes = sorted(s_set | c_set | e_set)
    rows = []
    for code in all_codes:
        in_s, in_c, in_e = code in s_set, code in c_set, code in e_set
        absent = [
            lbl for lbl, p in [("STIBO", in_s), ("CT", in_c), (erp_name, in_e)] if not p
        ]
        rows.append(
            {
                "Code": code,
                "STIBO": "X" if in_s else "",
                "CT": "X" if in_c else "",
                erp_name: "X" if in_e else "",
                "Absent_from": ", ".join(absent) if absent else "-",
            }
        )
    return {
        "rows": rows,
        "metrics": {
            "total": len(rows),
            "in_all_3": sum(1 for r in rows if r["Absent_from"] == "-"),
            "missing_stibo": sum(1 for r in rows if not r["STIBO"]),
            "missing_ct": sum(1 for r in rows if not r["CT"]),
            "missing_erp": sum(1 for r in rows if not r[erp_name]),
        },
    }


def reconcile_invoice_os(
    erp_name: str,
    ct_vendor_bytes: bytes | None,
    ct_customer_bytes: bytes | None,
    erp_vendor_bytes: bytes | None,
    erp_customer_bytes: bytes | None,
    stibo_vendor_invoice_bytes: bytes | None,
    stibo_vendor_os_bytes: bytes | None,
    stibo_customer_invoice_bytes: bytes | None,
    stibo_customer_os_bytes: bytes | None,
) -> dict:
    """Run Invoice + Ordering-Shipping reconciliation.

    Returns a dict with 'invoice' and 'ordering_shipping', each containing
    'vendor' and 'customer' sub-dicts with rows and metrics.
    """
    # Invoice
    stibo_v_inv = (
        load_stibo_vendor_invoice(stibo_vendor_invoice_bytes)
        if stibo_vendor_invoice_bytes
        else []
    )
    stibo_c_inv = (
        load_stibo_customer_invoice(stibo_customer_invoice_bytes)
        if stibo_customer_invoice_bytes
        else []
    )
    ct_v_inv = load_ct_vendor_invoice(ct_vendor_bytes) if ct_vendor_bytes else []
    ct_c_inv = load_ct_customer_invoice(ct_customer_bytes) if ct_customer_bytes else []
    erp_v_inv = (
        _load_erp_vendor_invoice(erp_vendor_bytes, erp_name) if erp_vendor_bytes else []
    )
    erp_c_inv = (
        _load_erp_customer_invoice(erp_customer_bytes, erp_name)
        if erp_customer_bytes
        else []
    )

    invoice = {
        "vendor": _build_entity_table(stibo_v_inv, ct_v_inv, erp_v_inv, erp_name),
        "customer": _build_entity_table(stibo_c_inv, ct_c_inv, erp_c_inv, erp_name),
    }

    # Ordering-Shipping (normalize codes to preserve leading zeros)
    stibo_v_os = _normalize_os(
        load_stibo_vendor_os(stibo_vendor_os_bytes) if stibo_vendor_os_bytes else []
    )
    stibo_c_os = _normalize_os(
        load_stibo_customer_os(stibo_customer_os_bytes)
        if stibo_customer_os_bytes
        else []
    )
    ct_v_os = _normalize_os(
        load_ct_vendor_os(ct_vendor_bytes) if ct_vendor_bytes else []
    )
    ct_c_os = _normalize_os(
        load_ct_customer_os(ct_customer_bytes) if ct_customer_bytes else []
    )
    erp_v_os = _normalize_os(
        _load_erp_vendor_os(erp_vendor_bytes, erp_name) if erp_vendor_bytes else []
    )
    erp_c_os = _normalize_os(
        _load_erp_customer_os(erp_customer_bytes, erp_name)
        if erp_customer_bytes
        else []
    )

    ordering_shipping = {
        "vendor": _build_entity_table(stibo_v_os, ct_v_os, erp_v_os, erp_name),
        "customer": _build_entity_table(stibo_c_os, ct_c_os, erp_c_os, erp_name),
    }

    return {
        "invoice": invoice,
        "ordering_shipping": ordering_shipping,
        "erp_name": erp_name,
    }
